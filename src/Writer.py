from openpyxl import Workbook
import openpyxl
import os
from datetime import datetime

class Writer():
    def __init__(self, PATH, filename):
        self.filename = filename
        self.PATH = PATH
        self.wb = None
        self.personal_headers = ['Invoice Reference', 'Client Name', 'Total Amount', 'Paid at Purchase', 'Paid at Credit', 'Payment Status', 'Date of Payment', 'Date of Purchase', 'Notes', 'Operator']
        self.bankers_headers = ['Invoice Reference', 'Insurance Reference', 'Group_Name', 'Client Name', 'Total Amount', 'Total Approved', 'Patient Share', 'Insurance Share', 'Insurance Fees', 'Payment Status', 'Date of Payment', 'Date of Purchase', 'Notes', 'Operator']
        self.globemed_headers = ['Invoice Reference', 'Client Name', 'Authorization Number', 'SSNBR', 'Total Approved', 'Net Amount', 'Patient Share', 'Insurance Share', 'Insurance Fees', 'Payment Status', 'Date of Payment', 'Date of Purchase', 'Notes', 'Operator']
        self.get_file(self.PATH)

    def get_file(self, PATH, filename="{}_{}.xlsx".format(datetime.now().month, datetime.now().year)): #Checks if the excel file is already available
        files = os.listdir(PATH)
        if filename in files:
            print('File already exists')
            self.wb = openpyxl.load_workbook(self.PATH + '\\' + filename)
        else:
            print('File does not exist yet')
            self.create_file(filename, self.PATH)

    def create_file(self, filename, PATH):
        #Create file
        self.wb = Workbook()
        self.wb.remove_sheet(self.wb["Sheet"])
        personal = self.wb.create_sheet('Personal',2)
        globemed = self.wb.create_sheet('Globemed',1)
        bankers = self.wb.create_sheet('Bankers',0)

        #Write headers
        self.write_headers('Bankers', self.bankers_headers)
        self.write_headers('Globemed', self.globemed_headers)
        self.write_headers('Personal', self.personal_headers)

        #Save Workbook
        self.wb.save(PATH + filename)

    def get_current_index(self,sheet_name):
        return   self.wb[sheet_name].max_row

    def get_current_filename(self):
        return "{}_{}.xlsx".format(datetime.now().month, datetime.now().year)

    def get_current_index(self, sheetname):
        return self.wb[sheetname].max_row

    def write_headers(self, sheet_name, headers):
        row_idx = self.get_current_index(sheet_name)
        for col_idx in range(1,len(headers)+1):
            self.wb[sheet_name].cell(row_idx, col_idx, headers[col_idx-1])
    
    def update_workbook(self, PATH=None, filename=None):
        if PATH != None:
            self.wb.save(PATH + '\\' + filename)
        else:
            self.wb.save(self.PATH + '\\' + self.filename)

    def write_data(self, sheet_name, data):
        row_idx = self.get_current_index(sheet_name)+1
        for col_idx in range(1,len(data)+1):
            self.wb[sheet_name].cell(row_idx, col_idx, data[col_idx-1])
        self.update_workbook()

if __name__ == "__main__":
    db = Writer()
    db.get_file(db.PATH)
    sample_bankers = [15487545, 115131215, 'LAU', 'Robin', 180000, 'None', 15000, 'None', 'None', 'CNS', str(datetime.now()), 'None', 'None', 'Dana'] #Bankers
    sample_globemed = [1516546, 'Ralph', 56164565, 154549978987, 15008551, 'None', 1500, 1515, 'None', 'D', 'None', str(datetime.now()), 'Notes', 'Roula']    
    db.write_data('Globemed', sample_globemed)
    db.write_data('Bankers', sample_bankers)
    db.update_workbook(db.PATH, db.filename)

